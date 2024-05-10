import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import pandas as pd
from datetime import date, datetime
from dateutil.relativedelta import relativedelta, MO, TH

# создаем книгу
wb = Workbook()
# делаем единственный лист активным
ws = wb.active

# указываем заглавие листа
ws.title = "Sheet"

# объединяем ячейки для указания заглавий
ws.merge_cells('C1:D1')
ws.merge_cells('G1:K1')
ws['C1'] = 'Старые'
ws['G1'] = 'Новые'

ws.merge_cells('B28:D28')
ws.merge_cells('G28:K28')
ws['B28'] = 'Добавить номера'
ws['G28'] = 'Удалить номера'

# расположим заглавия по центру объединенных ячеек

ws['C1'].alignment = Alignment(horizontal='center')
ws['G1'].alignment = Alignment(horizontal='center')
ws['B28'].alignment = Alignment(horizontal='center')
ws['G28'].alignment = Alignment(horizontal='center')

# указываем интересующую категорию
ws['A4'] = 'name_company1'
ws['A5'] = 'name_company2'
ws['A6'] = 'name_company3'
ws['A7'] = 'name_company4'
ws['A8'] = 'name_company5'

ws['A30'] = 'name_company1'
ws['A31'] = 'name_company2'
ws['A32'] = 'name_company3'
ws['A33'] = 'name_company4'
ws['A34'] = 'name_company5'

# открываем файл excel, с определенным названием, из которого, к примеру, будет извлекать абонентский номер с определенным статусом
data = pd.read_excel('./общий.xlsx')

# открываем еще один файл excel, откуда будет извлекать абонентские номера для сравнения, копирования
old_data = load_workbook('./локальный.xlsx').active
old_data.active = 1

# производим подсчет даты, в зависимости от дня недели. Нас интересует, исключительно, Понедельник и Четверг
date_now = datetime.now()
weekday_now = date_now.weekday()
if weekday_now in [3, 4, 5, 6]:
    remove_date = date_now + relativedelta(weekday=MO)
else:
    remove_date = date_now + relativedelta(weekday=TH)
# Форматируем дату в строку в формате "дд.мм.гггг"
remove_date = datetime.strftime(remove_date, '%Y-%m-%d')

number_list = data['number'].tolist()

in_use_name_company1 = data[data['agent'] == 'name_company1'][data['status'] == 'in use']['number'].tolist()
in_use_name_company2 = data[data['agent'] == 'name_company2'][data['status'] == 'in use']['number'].tolist()
in_use_name_company3 = data[data['agent'] == 'name_company3'][data['status'] == 'in use']['number'].tolist()
in_use_name_company4 = data[data['agent'] == 'name_company4'][data['status'] == 'in use']['number'].tolist()
in_use_name_company5 = data[data['agent'] == 'name_company5'][data['status'] == 'in use']['number'].tolist()

reserved_name_company1 = data[data['agent'] == 'name_company1'][data['status'] == 'reserved']['number'].tolist()
reserved_name_company2 = data[data['agent'] == 'name_company2'][data['status'] == 'reserved']['number'].tolist()
reserved_name_company3 = data[data['agent'] == 'name_company3'][data['status'] == 'reserved']['number'].tolist()
reserved_name_company4 = data[data['agent'] == 'name_company4'][data['status'] == 'reserved']['number'].tolist()
reserved_name_company5 = data[data['agent'] == 'name_company5'][data['status'] == 'reserved']['number'].tolist()

replaced_name_company1 = data[data['agent'] == 'name_company1'][data['status'] == 'replaced'][data['date remove'] == remove_date]['number'].tolist()
replaced_name_company2 = data[data['agent'] == 'name_company2'][data['status'] == 'replaced'][data['date remove'] == remove_date]['number'].tolist()
replaced_name_company3 = data[data['agent'] == 'name_company3'][data['status'] == 'replaced'][data['date remove'] == remove_date]['number'].tolist()
replaced_name_company4 = data[data['agent'] == 'name_company4'][data['status'] == 'replaced'][data['date remove'] == remove_date]['number'].tolist()
replaced_name_company5 = data[data['agent'] == 'name_company5'][data['status'] == 'replaced'][data['date remove'] == remove_date]['number'].tolist()

def find_no_spam_numbers(df1):
    no_spam_numbers = df1[df1['Спам'] == 'Нет']['Номер'].tolist()
    return no_spam_numbers

# Заполняем таблицу

# Заполняем ячейки номерами "Старые"
    # Копируем из старого списка старые номера
old_name_company1 = [old_data.cell(row=4, column=i).value for i in range(2, 6)]
old_name_company2 = [old_data.cell(row=5, column=i).value for i in range(2, 6)]
old_name_company3 = [old_data.cell(row=6, column=i).value for i in range(2, 6)]
old_name_company4 = [old_data.cell(row=7, column=i).value for i in range(2, 6)]
old_name_company5 = [old_data.cell(row=8, column=i).value for i in range(2, 6)]
    # Заполняем новую таблицу
for i, value in enumerate(old_name_company1, 1):
    ws.cell(column=i+1, row=4, value=value)
for i, value in enumerate(old_name_company2, 1):
    ws.cell(column=i+1, row=5, value=value)
for i, value in enumerate(old_name_company3, 1):
    ws.cell(column=i+1, row=6, value=value)
for i, value in enumerate(old_name_company4, 1):
    ws.cell(column=i+1, row=7, value=value)
for i, value in enumerate(old_name_company5, 1):
    ws.cell(column=i+1, row=8, value=value)

# Заполняем ячейки номерами "Новые"
for i, value in enumerate(in_use_name_company1, 6):
    ws.cell(column=i+1, row=4, value=value)
for i, value in enumerate(in_use_name_company2, 6):
    ws.cell(column=i+1, row=5, value=value)
for i, value in enumerate(in_use_name_company3, 6):
    ws.cell(column=i+1, row=6, value=value)
for i, value in enumerate(in_use_name_company4, 6):
    ws.cell(column=i+1, row=7, value=value)
for i, value in enumerate(in_use_name_company5, 6):
    ws.cell(column=i+1, row=8, value=value)

# Заполняем ячейки "Добавить номера"
for i, value in enumerate(reserved_name_company1, 1):
    ws.cell(column=i+1, row=30, value=value)
for i, value in enumerate(reserved_name_company2, 1):
    ws.cell(column=i+1, row=31, value=value)
for i, value in enumerate(reserved_name_company3, 1):
    ws.cell(column=i+1, row=32, value=value)
for i, value in enumerate(reserved_name_company4, 1):
    ws.cell(column=i+1, row=33, value=value)
for i, value in enumerate(reserved_name_company5, 1):
    ws.cell(column=i+1, row=34, value=value)

# Заполняем ячейки "Удалить номера"
for i, value in enumerate(replaced_name_company1, 6):
    ws.cell(column=i+1, row=30, value=value)
for i, value in enumerate(replaced_name_company2, 6):
    ws.cell(column=i+1, row=31, value=value)
for i, value in enumerate(replaced_name_company3, 6):
    ws.cell(column=i+1, row=32, value=value)
for i, value in enumerate(replaced_name_company4, 6):
    ws.cell(column=i+1, row=33, value=value)
for i, value in enumerate(replaced_name_company5, 6):
    ws.cell(column=i+1, row=34, value=value)

wb.save('./test.xlsx')





